import os
import boto3
from botocore.exceptions import ClientError
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


def upload_file_to_s3(file, form_id, field_name):
    """
    Upload file to S3 bucket and return the public URL.
    
    Args:
        file: UploadedFile object from request.FILES
        form_id: ID of the form
        field_name: Name of the field in the form schema
    
    Returns:
        str: Public URL of the uploaded file
    
    Raises:
        Exception: If upload fails
    """
    if not all([settings.AWS_ACCESS_KEY_ID, settings.AWS_SECRET_ACCESS_KEY, settings.AWS_STORAGE_BUCKET_NAME]):
        raise ValueError("AWS credentials are not configured. Please set AWS environment variables.")
    
    try:
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME
        )
        
        # Generate unique file name
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_extension = os.path.splitext(file.name)[1]
        s3_key = f'media/forms/{form_id}/{field_name}_{timestamp}{file_extension}'
        
        # Upload file without ACL (use bucket policy for public access instead)
        s3_client.upload_fileobj(
            file,
            settings.AWS_STORAGE_BUCKET_NAME,
            s3_key,
            ExtraArgs={
                'ContentType': file.content_type
            }
        )
        
        # Generate public URL
        file_url = f'https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{s3_key}'
        
        logger.info(f"File uploaded successfully to S3: {file_url}")
        return file_url
        
    except ClientError as e:
        logger.error(f"S3 upload failed: {str(e)}")
        raise Exception(f"Failed to upload file to S3: {str(e)}")
    except Exception as e:
        logger.error(f"Unexpected error during S3 upload: {str(e)}")
        raise


def generate_excel_export(form, responses):
    """
    Generate Excel file from form responses.
    
    Args:
        form: Form model instance
        responses: QuerySet of FormResponse objects
    
    Returns:
        BytesIO: Excel file as bytes
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Form Responses"
    
    # Get form schema
    schema = form.schema
    fields = schema.get('fields', [])
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create headers
    headers = ["Submission ID", "User Email", "Submitted At"]
    for field in fields:
        headers.append(field.get('name', 'Unknown'))
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        worksheet.column_dimensions[cell.column_letter].width = 20
    
    # Write data rows
    for row_idx, response in enumerate(responses, start=2):
        # Basic info
        worksheet.cell(row=row_idx, column=1).value = response.id
        worksheet.cell(row=row_idx, column=2).value = response.user.email
        worksheet.cell(row=row_idx, column=3).value = response.submitted_at.strftime('%Y-%m-%d %H:%M:%S')
        
        # Response data
        response_data = response.response_data
        for col_idx, field in enumerate(fields, start=4):
            field_name = field.get('name')
            field_value = response_data.get(field_name, '')
            
            # Handle file URLs
            if field.get('type') == 'file' and field_value:
                # Create hyperlink for file URLs
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = field_value
                cell.hyperlink = field_value
                cell.font = Font(color="0563C1", underline="single")
            else:
                worksheet.cell(row=row_idx, column=col_idx).value = str(field_value) if field_value else ''
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    # Save to BytesIO
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def has_file_fields(schema):
    """
    Check if form schema contains any file upload fields.
    
    Args:
        schema: Form schema dictionary
    
    Returns:
        bool: True if schema has file fields, False otherwise
    """
    fields = schema.get('fields', [])
    return any(field.get('type') == 'file' for field in fields)


def get_file_fields(schema):
    """
    Get list of file field names from schema.
    
    Args:
        schema: Form schema dictionary
    
    Returns:
        list: List of field names that are file type
    """
    fields = schema.get('fields', [])
    return [field.get('name') for field in fields if field.get('type') == 'file']
